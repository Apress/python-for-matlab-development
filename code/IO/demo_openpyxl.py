# code/IO/demo_openpyxl.py
# {{{
# This code accompanies the book _Python for MATLAB Development:
# Extend MATLAB with 300,000+ Modules from the Python Package Index_ 
# ISBN 978-1-4842-7222-0 | ISBN 978-1-4842-7223-7 (eBook)
# DOI 10.1007/978-1-4842-7223-7
# https://github.com/Apress/python-for-matlab-development
# 
# Copyright © 2022 Albert Danial
# 
# MIT License:
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# 
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
# 
# THE SOFTWARE IS PROVIDED “AS IS”, WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
# THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
# FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SOFTWARE.
# }}}
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
book = Workbook()
sheet = book.active
sheet.title = "Pets by weight"

# font styles, background color
ft_title = Font(name='Arial', size=14, bold=True)
ft_red   = Font(color='00FF0000')
ft_italics = Font(bold=True, italic=True)
bg_green = PatternFill(fgColor='C5FD2F', fill_type = 'solid')

sheet.merge_cells('B2:D3')
sheet['B2'] = 'My Pets'
sheet['B2'].font = ft_title
sheet['B2'].alignment = Alignment(
    horizontal="center", vertical="center")

# column headings
category = ['Name', 'Animal', 'weight [kg]']
row, col = 4, 2
for i in range(len(category)):
    cell = sheet.cell(row,col+i,category[i])
    cell.fill = bg_green
  
pets = [['Nutmeg', 'Rabbit', 2.5],
        ['Annabel', 'Dog', 4.3],
        ['Sunny', 'Bird', 0.02],
        ['Harley', 'Dog', 17.1],
        ['Toby', 'Dog', 24.0],
        ['Mr Socks', 'Cat', 3.9]]
for P in pets:
    row += 1
    for j in range(len(category)):
        cell = sheet.cell(row,col+j,P[j])
        if j == 2 and P[j] < 0.1:
            cell.font = ft_red

# equation to sum all weights
sheet[f'D{row+1}'] = f"=SUM(D4:D{row})"

row += 1
sheet.merge_cells(f'B{row}:C{row}')
sheet[f'B{row}'] = 'Total weight:'
sheet[f'B{row}'].font = ft_italics

book.save("pets.xlsx")
