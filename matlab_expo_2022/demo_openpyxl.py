#!/usr/bin/env python3

import openpyxl as OP
import openpyxl.styles as styles
Font = styles.Font
Alignment = styles.Alignment
PatternFill = styles.PatternFill
book = OP.Workbook()
sheet = book.active
sheet.title = "Pets by weight"

# font styles, background color
ft_title = Font(name="Arial",
                size=14,
                bold=True)
ft_red = Font(color="00FF0000")
ft_italics = Font(bold=True,
                  italic=True)
bg_green = PatternFill(
    fgColor="C5FD2F",fill_type="solid")

sheet.merge_cells("B2:D3")
B2 = sheet.cell(2,2)
B2.value = "My Pets"
B2.font = ft_title
B2.alignment = Alignment(
  horizontal="center",vertical="center")

# column headings
category=["Name","Animal","weight [kg]"]
row, col = 4, 2
for i in range(len(category)):
  nextCell = sheet.cell(row, col+i)
  nextCell.value = category[i]
  nextCell.fill = bg_green


pets = [["Nutmeg", "Rabbit", 2.5],
        ["Annabel", "Dog", 4.3],
        ["Sunny", "Bird", 0.02],
        ["Harley", "Dog", 17.1],
        ["Toby", "Dog", 24.0],
        ["Mr Socks", "Cat", 3.9]]

for P in pets:
  row += 1
  for j in range(len(category)):
    cell = sheet.cell(row,col+j,
                      P[j])
    if j == 2 and P[j] < 0.1:
      nextCell = sheet.cell(row,col+j)
      nextCell.font = ft_red




# equation to sum all weights
eqn = f"=SUM(D4:{row+1})"
nextCell = sheet.cell(row+1, 4)
nextCell.value = eqn

nextCell = sheet.cell(row+1, 2)
nextCell.value = "Total weight:"
nextCell.font = ft_italics;

book.save("pets.xlsx")
